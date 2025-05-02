import tkinter as tk
from typing import List, Dict, Optional
from datetime import date
from core.models import Revenu, Dépense, BudgetPrevisionnel
from core.data_handler import DataHandler
import openpyxl
from openpyxl.workbook import Workbook


class BudgetManager:
    def __init__(self):
        self.data_handler = DataHandler()
        self.revenus: List[Revenu] = []
        self.depenses: List[Dépense] = []
        self.budget_previsionnel: Optional[BudgetPrevisionnel] = None
        self._charger_les_donnees()

    def _charger_les_donnees(self):
        data = self.data_handler.charger_donnees()
        self.revenus = data["revenus"]
        self.depenses = data["depenses"]
        self.budget_previsionnel = data.get("budget_previsionnel")

    def _sauvegarder_les_donnees(self):
        data = {
            "revenus": self.revenus,
            "depenses": self.depenses,
            "budget_previsionnel": self.budget_previsionnel,
        }
        self.data_handler.sauvegarder_donnees(data)

    def ajouter_revenu(self, montant: float, source: str, date: date):
        revenu = Revenu(montant, source, date)
        self.revenus.append(revenu)
        self._sauvegarder_les_donnees()

    def ajouter_depense(self, montant: float, categorie: str, date: date, description: str = ""):
        depense = Dépense(montant, categorie, date, description)
        self.depenses.append(depense)
        self._sauvegarder_les_donnees()

    def calculer_solde(self) -> float:
        total_revenus = sum(revenu.montant for revenu in self.revenus)
        total_depenses = sum(depense.montant for depense in self.depenses)
        return total_revenus - total_depenses

    def get_all_revenus(self) -> List[Revenu]:
        return self.revenus

    def get_all_depenses(self) -> List[Dépense]:
        return self.depenses

    def definir_budget_previsionnel(self, montants_par_categorie: Dict[str, float]):
        self.budget_previsionnel = BudgetPrevisionnel(montants_par_categorie)
        self._sauvegarder_les_donnees()

    def get_budget_previsionnel(self) -> Optional[BudgetPrevisionnel]:
        return self.budget_previsionnel

    def calculer_depenses_par_categorie(self) -> Dict[str, float]:
        """Calcule le total des dépenses pour chaque catégorie."""
        depenses_par_categorie = {}
        for depense in self.depenses:
            categorie = depense.categorie
            montant = depense.montant
            if categorie in depenses_par_categorie:
                depenses_par_categorie[categorie] += montant
            else:
                depenses_par_categorie[categorie] = montant
        return depenses_par_categorie

    def calculer_depassements_budget(self) -> Dict[str, float]:
        """
        Calcule les dépassements de budget pour chaque catégorie.

        Retourne un dictionnaire où les clés sont les catégories et les valeurs sont les montants
        des dépassements (positifs si dépassement, zéro ou négatif si respect du budget).
        """
        depassements = {}
        if self.budget_previsionnel:
            depenses_par_categorie = self.calculer_depenses_par_categorie()
            for categorie, montant_budgetise in self.budget_previsionnel.montants_par_categorie.items():
                montant_reel = depenses_par_categorie.get(categorie, 0)  # 0 si la catégorie n'a pas de dépenses
                depassement = montant_reel - montant_budgetise
                depassements[categorie] = depassement
        return depassements

    def exporter_vers_excel(self, filename="budget_export.xlsx"):
        """
        Exporte les données (revenus, dépenses, budget prévisionnel) vers un fichier Excel.
        """
        workbook = Workbook()

        # Feuille pour les revenus
        sheet_revenus = workbook.active
        sheet_revenus.title = "Revenus"
        sheet_revenus.append(["Date", "Source", "Montant"])  # Ajoute l'en-tête
        for revenu in self.revenus:
            sheet_revenus.append([revenu.date, revenu.source, revenu.montant])

        # Feuille pour les dépenses
        sheet_depenses = workbook.create_sheet(title="Dépenses")
        sheet_depenses.append(["Date", "Catégorie", "Description", "Montant"])  # Ajoute l'en-tête
        for depense in self.depenses:
            sheet_depenses.append([depense.date, depense.categorie, depense.description, depense.montant])

        # Feuille pour le budget prévisionnel
        if self.budget_previsionnel:
            sheet_budget = workbook.create_sheet(title="Budget Prévisionnel")
            sheet_budget.append(["Catégorie", "Montant"])  # Ajoute l'en-tête
            for categorie, montant in self.budget_previsionnel.montants_par_categorie.items():
                sheet_budget.append([categorie, montant])

        try:
            workbook.save(filename)
            print(f"Données exportées avec succès vers {filename}")
        except Exception as e:
            print(f"Erreur lors de l'exportation vers Excel : {e}")

    def filtrer_revenus(self, date_min: date = None, date_max: date = None, source: str = None) -> List[Revenu]:
        """Filtre les revenus en fonction de la date et/ou de la source."""
        resultat = self.revenus
        if date_min:
            resultat = [revenu for revenu in resultat if revenu.date >= date_min]
        if date_max:
            resultat = [revenu for revenu in resultat if revenu.date <= date_max]
        if source:
            resultat = [revenu for revenu in resultat if source.lower() in revenu.source.lower()]
        return resultat

    def filtrer_depenses(self, date_min: date = None, date_max: date = None, categorie: str = None, description: str = None) -> List[Dépense]:
        """Filtre les dépenses en fonction de la date, de la catégorie et/ou de la description."""
        resultat = self.depenses
        if date_min:
            resultat = [depense for depense in resultat if depense.date >= date_min]
        if date_max:
            resultat = [depense for depense in resultat if depense.date <= date_max]
        if categorie:
            resultat = [depense for depense in resultat if categorie.lower() in depense.categorie.lower()]
        if description:
            resultat = [depense for depense in resultat if description.lower() in depense.description.lower()]
        return resultat